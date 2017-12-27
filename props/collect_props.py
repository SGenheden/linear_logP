"""
Collect molecular properties in an Excel sheet.
Calculates the properties with the Indigo package
"""
from __future__ import division, print_function, absolute_import

import sys
import urllib

import openpyxl as xl
from indigo import Indigo
ind = Indigo()

# Electronegativity on the Allen scale
electronegativity = {1:2.3 , 2:4.16 , 3:0.912 , 4:1.576 , 5:2.051 , 6:2.544 ,
                    7:3.066 , 8:3.61 , 9:4.193 , 10:4.787 , 11:0.869 , 12:1.293 ,
                    13:1.613 , 14:1.916 , 15:2.253 , 16:2.589 , 17:2.869 , 18:3.242 ,
                    19:0.734 , 20:1.034 , 21:1.19 , 22:1.38 , 23:1.53 , 24:1.65 ,
                    25:1.75 , 26:1.8 , 27:1.84 , 28:1.88 , 29:1.85 , 30:1.59 ,
                    31:1.756 , 32:1.994 , 33:2.211 , 34:2.424 , 35:2.685 , 36:2.966 ,
                    37:0.706 , 38:0.963 , 39:1.12 , 40:1.32 , 41:1.41 , 42:1.47 ,
                    43:1.51 , 44:1.54 , 45:1.56 , 46:1.58 , 47:1.87 , 48:1.52 ,
                    49:1.656 , 50:1.824 , 51:1.984 , 52:2.158 , 53:2.359 , 54:2.582 ,
                    55:0.659 , 56:0.881 , 71:1.09 , 72:1.16 , 73:1.34 , 74:1.47 ,
                    75:1.6 , 76:1.65 , 77:1.68 , 78:1.72 , 79:1.92 , 80:1.76 ,
                    81:1.789 , 82:1.854 , 83:2.01 , 84:2.19 , 85:2.39 , 86:2.6 , 87:0.67 , 88:0.89}

# Atomic masses
atomic_mass = [0.0, 1.0079, 4.0026, 6.941, 9.0122, 10.811, 12.0107, 14.0067, 15.9994, 18.9984,
        20.1797, 22.9898, 24.305, 26.9815, 28.0855, 30.9738, 32.065, 35.453, 39.948,
        39.0983, 40.078, 44.9559, 47.867, 50.9415, 51.9961, 54.938, 55.845, 58.9331,
        58.6934, 63.546, 65.409, 69.723, 72.64, 74.9216, 78.96, 79.904, 83.798, 85.4678,
        87.62, 88.9059, 91.224, 92.9064, 95.94, 98.0, 101.07, 102.9055, 106.42, 107.8682,
        112.411, 114.818, 118.71, 121.76, 127.6, 126.9045, 131.293, 132.9055,
        137.327, 138.9055, 140.116, 140.9077, 144.242, 145.0, 150.36, 151.964,
        157.25, 158.9254, 162.5, 164.9303, 167.259, 168.9342, 173.04, 174.967,
        178.49, 180.9479, 183.84, 186.207, 190.23, 192.217, 195.084, 196.9666,
        200.59, 204.3833, 207.2, 208.9804, 209.0, 210.0, 222.0, 223.0, 226.0,
        227.0, 232.0381, 231.0359, 238.0289, 237.0, 244.0, 243.0, 247.0, 247.0,
        251.0, 252.0]

def smiles2file(smiles) :
    """
    Convert SMILES using Cactus service
    Parameters
    ----------
    smiles : string
      the SMILES string
    Returns
    -------
    string
      the path to the downloaded SDF file
    Raises
    ------
    Exception
      if the retrieval fails
    """
    smiles = smiles.replace("#","%23")
    smiles = smiles.replace("+","%2B")
    site = "cactus.nci.nih.gov"
    try:
        page = urllib.urlopen("http://%s/cgi-bin/translate.tcl?smiles=%s&format=mol&astyle=kekule&dim=3D&file="%(site, smiles))
    except:
        raise Exception("Could not connect with server")

    for line in page:
        if "Click here" in line  and 'a href="' in line :
            dummy1, url, dummy2 = line.split('"')
            try:
                path, header = urllib.urlretrieve("http://%s%s"%(site,url))
            except:
                raise Exception("Could not retrieve file")
            return path

def gravitational_index(mol) :

    gi = 0
    for bond in mol.iterateBonds() :
        xyz1 = bond.source().xyz()
        xyz2 = bond.destination().xyz()
        d = (xyz1[0]-xyz2[0])**2 + (xyz1[1]-xyz2[1])**2 + (xyz1[2]-xyz2[2])**2
        gi += (atomic_mass[bond.source().atomicNumber()]* \
                atomic_mass[bond.destination().atomicNumber()]) / d
    return gi

def hbond_donors(mol) :

    q = ind.loadSmarts("[!$([#6,H0,-,-2,-3])]")
    try :
        return ind.substructureMatcher(mol).countMatches(q)
    except :
        return 0

def hbond_acceptors(mol) :

    q = ind.loadSmarts("[#6,#7;R0]=[#8]")
    try :
        return ind.substructureMatcher(mol).countMatches(q)
    except :
        return 0

def carbon_count(mol) :

    n = 0
    for atom in mol.iterateAtoms() :
        if atom.atomicNumber() == 6 :
            n +=1
    return n

def polar_count(mol) :

    carbon_en = electronegativity[6]
    n = 0
    for atom in mol.iterateAtoms() :
        try :
            if electronegativity[atom.atomicNumber()] > carbon_en :
                n +=1
        except :
            raise Exception("Unknown electronegativity=%d"%(atom.atomicNumber()))
    return n

wb = xl.load_workbook(filename = sys.argv[1])
ws = wb['data']

ws["E1"] = "Weight"
ws["F1"] = "Heavies"
ws["G1"] = "Carbons"
ws["H1"] = "Polar"
ws["I1"] = "Rings"
ws["J1"] = "HDonors"
ws["K1"] = "HAcceptors"
ws["L1"] = "GI"

i = 2
while True :
    smi = ws.cell(row=i, column=2).value
    if smi is None :
        break
    mol = ind.loadMolecule(smi)
    ws.cell(row=i, column=5).value = mol.molecularWeight()
    ws.cell(row=i, column=6).value = mol.countHeavyAtoms()
    ws.cell(row=i, column=7).value = carbon_count(mol)
    ws.cell(row=i, column=8).value = polar_count(mol)
    ws.cell(row=i, column=9).value = mol.countSSSR()
    ws.cell(row=i, column=10).value = hbond_donors(mol)
    ws.cell(row=i, column=11).value = hbond_acceptors(mol)


    for d in range(5) : # Try 5 times before giving up
        path = path = smiles2file(smi)
        if path is not None :
            break
    if path is not None :
        mol2 = ind.loadMoleculeFromFile(path)
        ws.cell(row=i, column=12).value = gravitational_index(mol2)
    else :
        ws.cell(row=i, column=12).value = -1
    i += 1

wb.save(sys.argv[1])
